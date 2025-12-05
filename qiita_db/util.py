r"""
Util functions (:mod: `qiita_db.util`)
======================================

..currentmodule:: qiita_db.util

This module provides different util functions.

Methods
-------

..autosummary::
    :toctree: generated/

    quote_data_value
    scrub_data
    exists_table
    get_db_files_base_dir
    compute_checksum
    get_files_from_uploads_folders
    filepath_id_to_rel_path
    filepath_id_to_object_id
    get_mountpoint
    insert_filepaths
    check_table_cols
    check_required_columns
    convert_from_id
    convert_to_id
    get_environmental_packages
    get_visibilities
    purge_filepaths
    move_filepaths_to_upload_folder
    move_upload_files_to_trash
    add_message
    get_pubmed_ids_from_dois
    generate_analysis_list
    human_merging_scheme
"""

# -----------------------------------------------------------------------------
# Copyright (c) 2014--, The Qiita Development Team.
#
# Distributed under the terms of the BSD 3-clause License.
#
# The full license is in the file LICENSE, distributed with this software.
# -----------------------------------------------------------------------------
import hashlib
from binascii import crc32
from contextlib import contextmanager
from csv import writer as csv_writer
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from errno import EEXIST
from functools import partial
from glob import glob
from io import StringIO
from itertools import chain
from json import loads
from os import listdir, makedirs, remove, stat, walk
from os.path import basename, exists, getsize, isdir, join
from random import SystemRandom
from shutil import copy as shutil_copy
from shutil import move, rmtree
from smtplib import SMTP, SMTP_SSL, SMTPException
from string import ascii_letters, digits, punctuation
from subprocess import check_output
from tempfile import mkstemp
from time import time as now

import h5py
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from bcrypt import gensalt, hashpw
from humanize import naturalsize
from matplotlib import colormaps
from openpyxl import load_workbook
from scipy.optimize import minimize

import qiita_db as qdb
from qiita_core.exceptions import IncompetentQiitaDeveloperError
from qiita_core.qiita_settings import qiita_config


def scrub_data(s):
    r"""Scrubs data fields of characters not allowed by PostgreSQL

    disallowed characters:
        '   ;

    Parameters
    ----------
    s : str
        The string to clean up

    Returns
    -------
    str
        The scrubbed string
    """
    ret = s.replace("'", "")
    ret = ret.replace(";", "")
    return ret


def convert_type(obj):
    """Converts a passed item to int, float, or str in that order

    Parameters
    ----------
    obj : object
        object to evaluate

    Returns
    -------
    int, float, or str
        Re-typed information from obj

    Raises
    ------
    IncompetentQiitaDeveloperError
        If the object can't be converted to int, float, or string

    Notes
    -----
    The function first tries to convert to an int. If that fails, it tries to
    convert to a float. If that fails it returns the original string.
    """
    item = None
    if isinstance(obj, datetime):
        item = str(obj)
    else:
        for fn in (int, float, str):
            try:
                item = fn(obj)
            except ValueError:
                continue
            else:
                break
    if item is None:
        raise IncompetentQiitaDeveloperError(
            "Can't convert item of type %s!" % str(type(obj))
        )
    return item


def get_artifact_types(key_by_id=False):
    """Gets the list of possible artifact types

    Parameters
    ----------
    key : bool, optional
        Determines the format of the returned dict. Defaults to false.

    Returns
    -------
    dict
        If key_by_id is True, dict is of the form
        {artifact_type_id: artifact_type}
        If key_by_id is False, dict is of the form
        {artifact_type: artifact_type_id}
    """
    with qdb.sql_connection.TRN:
        cols = (
            "artifact_type_id, artifact_type"
            if key_by_id
            else "artifact_type, artifact_type_id"
        )
        sql = "SELECT {} FROM qiita.artifact_type".format(cols)
        qdb.sql_connection.TRN.add(sql)
        return dict(qdb.sql_connection.TRN.execute_fetchindex())


def get_filepath_types(key="filepath_type"):
    """Gets the list of possible filepath types from the filetype table

    Parameters
    ----------
    key : {'filepath_type', 'filepath_type_id'}, optional
        Defaults to "filepath_type". Determines the format of the returned
        dict.

    Returns
    -------
    dict
        - If `key` is "filepath_type", dict is of the form
          {filepath_type: filepath_type_id}
        - If `key` is "filepath_type_id", dict is of the form
          {filepath_type_id: filepath_type}
    """
    with qdb.sql_connection.TRN:
        if key == "filepath_type":
            cols = "filepath_type, filepath_type_id"
        elif key == "filepath_type_id":
            cols = "filepath_type_id, filepath_type"
        else:
            raise qdb.exceptions.QiitaDBColumnError(
                "Unknown key. Pass either 'filepath_type' or 'filepath_type_id'."
            )
        sql = "SELECT {} FROM qiita.filepath_type".format(cols)
        qdb.sql_connection.TRN.add(sql)
        return dict(qdb.sql_connection.TRN.execute_fetchindex())


def get_data_types(key="data_type"):
    """Gets the list of possible data types from the data_type table

    Parameters
    ----------
    key : {'data_type', 'data_type_id'}, optional
        Defaults to "data_type". Determines the format of the returned dict.

    Returns
    -------
    dict
        - If `key` is "data_type", dict is of the form
          {data_type: data_type_id}
        - If `key` is "data_type_id", dict is of the form
          {data_type_id: data_type}
    """
    with qdb.sql_connection.TRN:
        if key == "data_type":
            cols = "data_type, data_type_id"
        elif key == "data_type_id":
            cols = "data_type_id, data_type"
        else:
            raise qdb.exceptions.QiitaDBColumnError(
                "Unknown key. Pass either 'data_type_id' or 'data_type'."
            )
        sql = "SELECT {} FROM qiita.data_type".format(cols)
        qdb.sql_connection.TRN.add(sql)
        return dict(qdb.sql_connection.TRN.execute_fetchindex())


def create_rand_string(length, punct=True):
    """Returns a string of random ascii characters

    Parameters
    ----------
    length: int
        Length of string to return
    punct: bool, optional
        Include punctuation as well as letters and numbers. Default True.
    """
    chars = ascii_letters + digits
    if punct:
        chars += punctuation
    sr = SystemRandom()
    return "".join(sr.choice(chars) for i in range(length))


def hash_password(password, hashedpw=None):
    """Hashes password

    Parameters
    ----------
    password: str
        Plaintext password
    hashedpw: str, optional
        Previously hashed password for bcrypt to pull salt from. If not
        given, salt generated before hash

    Returns
    -------
    str
        Hashed password

    Notes
    -----
    Relies on bcrypt library to hash passwords, which stores the salt as
    part of the hashed password. Don't need to actually store the salt
    because of this.
    """
    # all the encode/decode as a python 3 workaround for bcrypt
    if hashedpw is None:
        hashedpw = gensalt()
    else:
        hashedpw = hashedpw.encode("utf-8")
    password = password.encode("utf-8")
    output = hashpw(password, hashedpw)
    if isinstance(output, bytes):
        output = output.decode("utf-8")
    return output


def check_required_columns(keys, table):
    """Makes sure all required columns in database table are in keys

    Parameters
    ----------
    keys: iterable
        Holds the keys in the dictionary
    table: str
        name of the table to check required columns

    Raises
    ------
    QiitaDBColumnError
        If keys exist that are not in the table
    RuntimeError
        Unable to get columns from database
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT is_nullable, column_name, column_default
                 FROM information_schema.columns WHERE table_name = %s"""
        qdb.sql_connection.TRN.add(sql, [table])
        cols = qdb.sql_connection.TRN.execute_fetchindex()
        # Test needed because a user with certain permissions can query without
        # error but be unable to get the column names
        if len(cols) == 0:
            raise RuntimeError("Unable to fetch column names for table %s" % table)
        required = set(x[1] for x in cols if x[0] == "NO" and x[2] is None)
        if len(required.difference(keys)) > 0:
            raise qdb.exceptions.QiitaDBColumnError(
                "Required keys missing: %s" % required.difference(keys)
            )


def check_table_cols(keys, table):
    """Makes sure all keys correspond to column headers in a table

    Parameters
    ----------
    keys: iterable
        Holds the keys in the dictionary
    table: str
        name of the table to check column names

    Raises
    ------
    QiitaDBColumnError
        If a key is found that is not in table columns
    RuntimeError
        Unable to get columns from database
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT column_name FROM information_schema.columns
                 WHERE table_name = %s"""
        qdb.sql_connection.TRN.add(sql, [table])
        cols = qdb.sql_connection.TRN.execute_fetchflatten()
        # Test needed because a user with certain permissions can query without
        # error but be unable to get the column names
        if len(cols) == 0:
            raise RuntimeError("Unable to fetch column names for table %s" % table)
        if len(set(keys).difference(cols)) > 0:
            raise qdb.exceptions.QiitaDBColumnError(
                "Non-database keys found: %s" % set(keys).difference(cols)
            )


def get_table_cols(table):
    """Returns the column headers of table

    Parameters
    ----------
    table : str
        The table name

    Returns
    -------
    list of str
        The column headers of `table`
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT column_name FROM information_schema.columns
                 WHERE table_name=%s AND table_schema='qiita'"""
        qdb.sql_connection.TRN.add(sql, [table])
        return qdb.sql_connection.TRN.execute_fetchflatten()


def exists_table(table):
    r"""Checks if `table` exists on the database

    Parameters
    ----------
    table : str
        The table name to check if exists

    Returns
    -------
    bool
        Whether `table` exists on the database or not
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT exists(
                    SELECT table_name FROM information_schema.tables
                    WHERE table_name=%s)"""
        qdb.sql_connection.TRN.add(sql, [table])
        return qdb.sql_connection.TRN.execute_fetchlast()


def get_db_files_base_dir():
    r"""Returns the path to the base directory of all db files

    Returns
    -------
    str
        The path to the base directory of all db files
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT base_data_dir FROM settings")
        basedir = qdb.sql_connection.TRN.execute_fetchlast()
        # making sure that it never ends in a "/" as most tests expect this
        if basedir.endswith("/"):
            basedir = basedir[:-1]
        return basedir


def get_work_base_dir():
    r"""Returns the path to the base directory of all db files

    Returns
    -------
    str
        The path to the base directory of all db files
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT base_work_dir FROM settings")
        return qdb.sql_connection.TRN.execute_fetchlast()


def max_preparation_samples():
    r"""Returns the max number of samples allowed in a single preparation

    Returns
    -------
    int
        The max number of samples allowed in a single preparation
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT max_preparation_samples FROM settings")
        return qdb.sql_connection.TRN.execute_fetchlast()


def max_artifacts_in_workflow():
    r"""Returns the max number of artifacts allowed in a single workflow

    Returns
    -------
    int
        The max number of artifacts allowed in a single workflow
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT max_artifacts_in_workflow FROM settings")
        return qdb.sql_connection.TRN.execute_fetchlast()


def compute_checksum(path):
    r"""Returns the checksum of the file pointed by path

    Parameters
    ----------
    path : str
        The path to compute the checksum

    Returns
    -------
    int
        The file checksum
    """
    filepaths = []
    if isdir(path):
        for name, dirs, files in walk(path):
            join_f = partial(join, name)
            filepaths.extend(list(map(join_f, files)))
    else:
        filepaths.append(path)

    buffersize = 65536
    crcvalue = 0
    for fp in filepaths:
        with open(fp, "rb") as f:
            buffr = f.read(buffersize)
            while len(buffr) > 0:
                crcvalue = crc32(buffr, crcvalue)
                buffr = f.read(buffersize)
    # We need the & 0xFFFFFFFF in order to get the same numeric value across
    # all python versions and platforms
    return crcvalue & 0xFFFFFFFF


def get_files_from_uploads_folders(study_id):
    """Retrieve files in upload folders

    Parameters
    ----------
    study_id : str
        The study id of which to retrieve all upload folders

    Returns
    -------
    list
        List of the filepaths for upload for that study
    """
    study_id = str(study_id)
    fp = []
    for pid, p in get_mountpoint("uploads", retrieve_all=True):
        t = join(p, study_id)
        if exists(t):
            for f in listdir(t):
                d = join(t, f)
                if not f.startswith(".") and not isdir(d):
                    fp.append((pid, f, naturalsize(getsize(d), gnu=True)))

    return fp


def move_upload_files_to_trash(study_id, files_to_move):
    """Move files to a trash folder within the study_id upload folder

    Parameters
    ----------
    study_id : int
        The study id
    files_to_move : list
        List of tuples (folder_id, filename)

    Raises
    ------
    QiitaDBError
        If folder_id or the study folder don't exist and if the filename to
        erase matches the trash_folder, internal variable
    """
    trash_folder = "trash"
    folders = {k: v for k, v in get_mountpoint("uploads", retrieve_all=True)}

    for fid, filename in files_to_move:
        if filename == trash_folder:
            raise qdb.exceptions.QiitaDBError(
                "You can not erase the trash folder: %s" % trash_folder
            )

        if fid not in folders:
            raise qdb.exceptions.QiitaDBError(
                "The filepath id: %d doesn't exist in the database" % fid
            )

        foldername = join(folders[fid], str(study_id))
        if not exists(foldername):
            raise qdb.exceptions.QiitaDBError(
                "The upload folder for study id: %d doesn't exist" % study_id
            )

        trashpath = join(foldername, trash_folder)
        create_nested_path(trashpath)

        fullpath = join(foldername, filename)
        new_fullpath = join(foldername, trash_folder, filename)

        if exists(fullpath):
            move(fullpath, new_fullpath)


def get_mountpoint(mount_type, retrieve_all=False, retrieve_subdir=False):
    r"""Returns the most recent values from data directory for the given type

    Parameters
    ----------
    mount_type : str
        The data mount type
    retrieve_all : bool, optional
        Retrieve all the available mount points or just the active one.
        Default: False.
    retrieve_subdir : bool, optional
        Retrieve the subdirectory column. Default: False.

    Returns
    -------
    list
        List of tuple, where: [(id_mountpoint, filepath_of_mountpoint)]
    """
    with qdb.sql_connection.TRN:
        if retrieve_all:
            sql = """SELECT data_directory_id, mountpoint, subdirectory
                     FROM qiita.data_directory
                     WHERE data_type=%s ORDER BY active DESC"""
        else:
            sql = """SELECT data_directory_id, mountpoint, subdirectory
                     FROM qiita.data_directory
                     WHERE data_type=%s AND active=true"""
        qdb.sql_connection.TRN.add(sql, [mount_type])
        db_result = qdb.sql_connection.TRN.execute_fetchindex()
        basedir = get_db_files_base_dir()
        if retrieve_subdir:
            result = [(d, join(basedir, m), s) for d, m, s in db_result]
        else:
            result = [(d, join(basedir, m)) for d, m, _ in db_result]
        return result


def get_mountpoint_path_by_id(mount_id):
    r"""Returns the mountpoint path for the mountpoint with id = mount_id

    Parameters
    ----------
    mount_id : int
        The mountpoint id

    Returns
    -------
    str
        The mountpoint path
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT mountpoint FROM qiita.data_directory
                 WHERE data_directory_id=%s"""
        qdb.sql_connection.TRN.add(sql, [mount_id])
        mountpoint = qdb.sql_connection.TRN.execute_fetchlast()
        return join(get_db_files_base_dir(), mountpoint)


def insert_filepaths(filepaths, obj_id, table, move_files=True, copy=False):
    r"""Inserts `filepaths` in the database.

    Since the files live outside the database, the directory in which the files
    lives is controlled by the database, so it moves the filepaths from
    its original location to the controlled directory.

    Parameters
    ----------
    filepaths : iterable of tuples (str, int)
        The list of paths to the raw files and its filepath type identifier
    obj_id : int
        Id of the object calling the functions. Disregarded if move_files
        is False
    table : str
        Table that holds the file data
    move_files : bool, optional
        Whether or not to move the given filepaths to the db filepaths
        default: True
    copy : bool, optional
        If `move_files` is true, whether to actually move the files or just
        copy them

    Returns
    -------
    list of int
        List of the filepath_id in the database for each added filepath
    """
    with qdb.sql_connection.TRN:
        new_filepaths = filepaths

        dd_id, mp, subdir = get_mountpoint(table, retrieve_subdir=True)[0]
        base_fp = join(get_db_files_base_dir(), mp)

        if move_files or copy:
            db_path = partial(join, base_fp)
            if subdir:
                # Generate the new filepaths, format:
                # mountpoint/obj_id/original_name
                dirname = db_path(str(obj_id))
                create_nested_path(dirname)
                new_filepaths = [
                    (join(dirname, basename(path)), id_) for path, id_ in filepaths
                ]
            else:
                # Generate the new fileapths. format:
                # mountpoint/DataId_OriginalName
                new_filepaths = [
                    (db_path("%s_%s" % (obj_id, basename(path))), id_)
                    for path, id_ in filepaths
                ]
            # Move the original files to the controlled DB directory
            transfer_function = shutil_copy if copy else move
            for old_fp, new_fp in zip(filepaths, new_filepaths):
                transfer_function(old_fp[0], new_fp[0])
                # In case the transaction executes a rollback, we need to
                # make sure the files have not been moved
                qdb.sql_connection.TRN.add_post_rollback_func(
                    move, new_fp[0], old_fp[0]
                )

        def str_to_id(x):
            return x if isinstance(x, int) else convert_to_id(x, "filepath_type")

        # 1 is the checksum algorithm, which we only have one implemented
        values = [
            [
                basename(path),
                str_to_id(id_),
                compute_checksum(path),
                getsize(path),
                1,
                dd_id,
            ]
            for path, id_ in new_filepaths
        ]
        # Insert all the filepaths at once and get the filepath_id back
        sql = """INSERT INTO qiita.filepath
                    (filepath, filepath_type_id, checksum, fp_size,
                     checksum_algorithm_id, data_directory_id)
                 VALUES (%s, %s, %s, %s, %s, %s)
                 RETURNING filepath_id"""
        idx = qdb.sql_connection.TRN.index
        qdb.sql_connection.TRN.add(sql, values, many=True)
        # Since we added the query with many=True, we've added len(values)
        # queries to the transaction, so the ids are in the last idx queries
        return list(
            chain.from_iterable(
                chain.from_iterable(qdb.sql_connection.TRN.execute()[idx:])
            )
        )


def _path_builder(db_dir, filepath, mountpoint, subdirectory, obj_id):
    """Builds the path of a DB stored file

    Parameters
    ----------
    db_dir : str
        The DB base dir
    filepath : str
        The path stored in the DB
    mountpoint : str
        The mountpoint of the given file
    subdirectory : bool
        Whether the file is stored in a subdirectory in the mountpoint or not
    obj_id : int
        The id of the object to which the file is attached

    Returns
    -------
    str
        The full path of the given file
    """
    if subdirectory:
        return join(db_dir, mountpoint, str(obj_id), filepath)
    else:
        return join(db_dir, mountpoint, filepath)


def retrieve_filepaths(obj_fp_table, obj_id_column, obj_id, sort=None, fp_type=None):
    """Retrieves the filepaths for the given object id

    Parameters
    ----------
    obj_fp_table : str
        The name of the table that links the object and the filepath
    obj_id_column : str
        The name of the column that represents the object id
    obj_id : int
        The object id
    sort : {'ascending', 'descending'}, optional
        The direction in which the results are sorted, using the filepath id
        as sorting key. Default: None, no sorting is applied
    fp_type: str, optional
        Retrieve only the filepaths of the matching filepath type

    Returns
    -------
    list of dict {fp_id, fp, ft_type, checksum, fp_size}
        The list of dict with the properties of the filepaths
    """

    sql_sort = ""
    if sort == "ascending":
        sql_sort = " ORDER BY filepath_id"
    elif sort == "descending":
        sql_sort = " ORDER BY filepath_id DESC"
    elif sort is not None:
        raise qdb.exceptions.QiitaDBError(
            "Unknown sorting direction: %s. Please choose from 'ascending' or "
            "'descending'" % sort
        )

    sql_args = [obj_id]

    sql_type = ""
    if fp_type:
        sql_type = " AND filepath_type=%s"
        sql_args.append(fp_type)

    with qdb.sql_connection.TRN:
        sql = """SELECT filepath_id, filepath, filepath_type, mountpoint,
                        subdirectory, checksum, fp_size
                 FROM qiita.filepath
                    JOIN qiita.filepath_type USING (filepath_type_id)
                    JOIN qiita.data_directory USING (data_directory_id)
                    JOIN qiita.{0} USING (filepath_id)
                 WHERE {1} = %s{2}{3}""".format(
            obj_fp_table, obj_id_column, sql_type, sql_sort
        )
        qdb.sql_connection.TRN.add(sql, sql_args)
        results = qdb.sql_connection.TRN.execute_fetchindex()
        db_dir = get_db_files_base_dir()

        return [
            {
                "fp_id": fpid,
                "fp": _path_builder(db_dir, fp, m, s, obj_id),
                "fp_type": fp_type_,
                "checksum": c,
                "fp_size": fpsize,
            }
            for fpid, fp, fp_type_, m, s, c, fpsize in results
        ]


def _rm_files(TRN, fp):
    # Remove the data
    if exists(fp):
        if isdir(fp):
            func = rmtree
        else:
            func = remove
        TRN.add_post_commit_func(func, fp)


def purge_filepaths(delete_files=True):
    r"""Goes over the filepath table and removes all the filepaths that are not
    used in any place

    Parameters
    ----------
    delete_files : bool
        if True it will actually delete the files, if False print
    """
    with qdb.sql_connection.TRN:
        files_to_remove = []
        # qiita can basically download 5 things: references, info files,
        # artifacts, analyses & working_dir.
        # 1. references are not longer used so we can skip

        # 2. info files: here we could remove all old info files (the backup we
        #    keep when a user uploads a new file) and all info files from
        #    studies that no longer exist. We want to keep the old templates
        #    so we can recover them (this has happened before) but let's remove
        #    those from deleted studies. Note that we need to check for sample,
        #    prep and qiime info files
        st_id = qdb.util.convert_to_id("sample_template", "filepath_type")
        pt_id = qdb.util.convert_to_id("prep_template", "filepath_type")
        qt_id = qdb.util.convert_to_id("qiime_map", "filepath_type")
        sql = """SELECT filepath_id, filepath FROM qiita.filepath
                 WHERE filepath_type_id IN %s AND filepath ~ '^[0-9]' AND
                    data_directory_id = %s AND filepath_id NOT IN (
                        SELECT filepath_id FROM qiita.prep_template_filepath
                        UNION
                        SELECT filepath_id FROM qiita.sample_template_filepath)
              """
        for mp_id, mp in get_mountpoint("templates"):
            qdb.sql_connection.TRN.add(sql, [tuple([st_id, pt_id, qt_id]), mp_id])
            studies_exits = []
            studies_erased = []
            for fid, fp in qdb.sql_connection.TRN.execute_fetchindex():
                # making sure the studies do _not_ exist, remember info files
                # are prepended by the study id
                study_id = int(fp.split("_")[0])
                if study_id in studies_exits:
                    continue
                elif study_id in studies_erased:
                    fpath = qdb.util.get_filepath_information(fid)["fullpath"]
                    files_to_remove.append([fid, fpath])
                else:
                    try:
                        qdb.study.Study(study_id)
                    except qdb.exceptions.QiitaDBUnknownIDError:
                        fpath = qdb.util.get_filepath_information(fid)["fullpath"]
                        files_to_remove.append([fid, fpath])
                        studies_erased.append(study_id)
                    else:
                        studies_exits.append(study_id)

        # 3. artifacts: [A] the difficulty of deleting artifacts is that (1)
        #    they live in different mounts, (2) as inidividual folders [the
        #    artifact id], (3) and the artifact id within the database has
        #    been lost. Thus, the easiest is to loop over the different data
        #    directories (mounts), get the folder names (artifact ids), and
        #    check if they exist; if they don't let's delete them. [B] As an
        #    additional and final step, we need to purge these filepaths from
        #    the DB.
        #    [A]
        main_sql = """SELECT data_directory_id FROM qiita.artifact_type at
                        LEFT JOIN qiita.data_directory dd ON (
                            dd.data_type = at.artifact_type)
                        WHERE subdirectory = true"""
        qdb.sql_connection.TRN.add(main_sql)
        for mp_id in qdb.sql_connection.TRN.execute_fetchflatten():
            mount = get_mountpoint_path_by_id(mp_id)
            for fpath in listdir(mount):
                full_fpath = join(mount, fpath)
                if isdir(full_fpath):
                    try:
                        qdb.artifact.Artifact(int(fpath))
                    except qdb.exceptions.QiitaDBUnknownIDError:
                        files_to_remove.append([None, full_fpath])
                    else:
                        continue
        #    [B]
        sql = """SELECT filepath_id FROM qiita.filepath
                 WHERE filepath_id not in (
                    SELECT filepath_id FROM qiita.artifact_filepath) AND
                data_directory_id in (
                    SELECT data_directory_id FROM qiita.artifact_type at
                        LEFT JOIN qiita.data_directory dd ON (
                            dd.data_type = at.artifact_type)
                    WHERE subdirectory = true)
              """
        qdb.sql_connection.TRN.add(sql)
        for fid in qdb.sql_connection.TRN.execute_fetchflatten():
            fpath = qdb.util.get_filepath_information(fid)["fullpath"]
            aid = fpath.split("/")[-2]
            # making sure the artifact doesn't exist any more
            if aid == "None":
                files_to_remove.append([fid, None])

        # 4. analysis: we need to select all the filepaths that are not in
        #    the analysis_filepath, this will return both all filepaths not
        #    from analyses and those that are not being used, thus, we need
        #    to also not select those files that are not part of the artifacts
        #    by ignoring those files paths not stored in a data_directory from
        #    an artifact:
        sql = """SELECT filepath_id FROM qiita.filepath
                 WHERE filepath_id not in (
                    SELECT filepath_id FROM qiita.analysis_filepath) AND
                data_directory_id in (
                    SELECT data_directory_id FROM qiita.data_directory
                    WHERE data_type = 'analysis')
              """
        qdb.sql_connection.TRN.add(sql)
        for fid in qdb.sql_connection.TRN.execute_fetchflatten():
            fdata = qdb.util.get_filepath_information(fid)
            analysis_id = int(fdata["filepath"].split("_")[0])
            # making sure the Analysis doesn't exist
            if not qdb.analysis.Analysis.exists(analysis_id):
                fpath = fdata["fullpath"]
                files_to_remove.append([fid, fpath])

        # 5. working directory: this is done internally in the Qiita system via
        #    a cron job

        # Deleting the files!
        sql = "DELETE FROM qiita.filepath WHERE filepath_id = %s"
        for fid, fpath in files_to_remove:
            if delete_files:
                if fid is not None:
                    qdb.sql_connection.TRN.add(sql, [fid])
                if fpath is not None:
                    _rm_files(qdb.sql_connection.TRN, fpath)
            else:
                print("%s: %s" % (fid, fpath))

        if delete_files:
            # there is a chance that we will never enter the above
            # "if fid is not None" statement so we will add an extra SQL
            # command just to make sure that something gets executed
            qdb.sql_connection.TRN.add("SELECT 42")

            qdb.sql_connection.TRN.execute()


def quick_mounts_purge():
    r"""This is a quick mount purge as it only slightly relies on the database

    Notes
    -----
        Currently we delete anything older than 30 days that is not linked
        to the database. This number is intentionally hardcoded in the code.
        At the time of this writing this number seem high but keeping it
        this way to be safe. In the future, if needed, it can be changed.
    """
    with qdb.sql_connection.TRN:
        main_sql = """SELECT data_directory_id FROM qiita.artifact_type at
                  LEFT JOIN qiita.data_directory dd ON (
                      dd.data_type = at.artifact_type)
                  WHERE subdirectory = true"""
        qdb.sql_connection.TRN.add(main_sql)
        mp_ids = qdb.sql_connection.TRN.execute_fetchflatten()
        mounts = [qdb.util.get_mountpoint_path_by_id(x) for x in mp_ids]
        folders = [join(x, f) for x in mounts for f in listdir(x) if f.isnumeric()]

    # getting all unlinked folders
    to_delete = []
    for i, f in enumerate(folders):
        vals = f.split("/")
        aid = int(vals[-1])
        artifact_type = vals[-2]
        if artifact_type == "FeatureData[Taxonomy]":
            continue

        try:
            a = qdb.artifact.Artifact(aid)
        except qdb.exceptions.QiitaDBUnknownIDError:
            to_delete.append(f)
            continue
        if not a.artifact_type.startswith(artifact_type):
            raise ValueError(
                f"Review artifact type: {a.id} {artifact_type} {a.artifact_type}"
            )

    # now, let's just keep those older than 30 days (in seconds)
    ignore = now() - (30 * 86400)
    to_keep = [x for x in to_delete if stat(x).st_mtime >= ignore]
    to_delete = set(to_delete) - set(to_keep)

    # get stats to report
    stats = dict()
    for td in to_delete:
        f = td.split("/")[-2]
        if f not in stats:
            stats[f] = 0
        stats[f] += sum([getsize(join(p, fp)) for p, ds, fs in walk(td) for fp in fs])

    report = ["----------------------"]
    for f, s in stats.items():
        report.append(f"{f}\t{naturalsize(s)}")
    report.append(f"Total files {len(to_delete)} {naturalsize(sum(stats.values()))}")
    report.append("----------------------")

    for td in list(to_delete):
        if exists(td):
            rmtree(td)

    return "\n".join(report)


def _rm_exists(fp, obj, _id, delete_files):
    try:
        _id = int(_id)
        obj(_id)
    except Exception:
        _id = str(_id)
        if delete_files:
            with qdb.sql_connection.TRN:
                _rm_files(qdb.sql_connection.TRN, fp)
                qdb.sql_connection.TRN.execute()
        else:
            print("Remove %s" % fp)


def empty_trash_upload_folder(delete_files=True):
    r"""Delete all files in the trash folder inside each of the upload
    folders

    Parameters
    ----------
    delete_files : bool
        if True it will actually delete the files, if False print
    """
    gfp = partial(join, get_db_files_base_dir())
    with qdb.sql_connection.TRN:
        sql = """SELECT mountpoint
                 FROM qiita.data_directory
                 WHERE data_type = 'uploads'"""
        qdb.sql_connection.TRN.add(sql)

        for mp in qdb.sql_connection.TRN.execute_fetchflatten():
            for path, dirs, files in walk(gfp(mp)):
                if path.endswith("/trash"):
                    if delete_files:
                        for f in files:
                            fp = join(path, f)
                            _rm_files(qdb.sql_connection.TRN, fp)
                    else:
                        print(files)

        if delete_files:
            qdb.sql_connection.TRN.execute()


def move_filepaths_to_upload_folder(study_id, filepaths):
    r"""Goes over the filepaths list and moves all the filepaths that are not
    used in any place to the upload folder of the study

    Parameters
    ----------
    study_id : int
        The study id to where the files should be returned to
    filepaths : list
        List of filepaths to move to the upload folder
    """
    with qdb.sql_connection.TRN:
        uploads_fp = join(get_mountpoint("uploads")[0][1], str(study_id))

        create_nested_path(uploads_fp)

        path_builder = partial(join, uploads_fp)

        # do not move these files-types back to upload folder.
        do_not_move = [
            "preprocessed_fasta",
            "preprocessed_fastq",
            "preprocessed_demux",
            "directory",
            "log",
            "html_summary",
            "tgz",
            "html_summary_dir",
            "qzv",
            "qza",
        ]

        # We can now go over and remove all the filepaths
        sql = """DELETE FROM qiita.filepath WHERE filepath_id = %s"""
        for x in filepaths:
            qdb.sql_connection.TRN.add(sql, [x["fp_id"]])

            if x["fp_type"] in do_not_move:
                _rm_files(qdb.sql_connection.TRN, x["fp"])
                continue

            # if files were not removed, then they should be moved.
            destination = path_builder(basename(x["fp"]))
            qdb.sql_connection.TRN.add_post_rollback_func(move, destination, x["fp"])
            move(x["fp"], destination)

        qdb.sql_connection.TRN.execute()


def get_filepath_information(filepath_id):
    """Gets the filepath information of filepath_id

    Parameters
    ----------
    filepath_id : int
        The filepath id

    Returns
    -------
    dict
        The filepath information
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT filepath_id, filepath, filepath_type, checksum,
                        data_type, mountpoint, subdirectory, active,
                        artifact_id
                 FROM qiita.filepath
                    JOIN qiita.filepath_type USING (filepath_type_id)
                    JOIN qiita.data_directory USING (data_directory_id)
                    LEFT JOIN qiita.artifact_filepath USING (filepath_id)
                 WHERE filepath_id = %s"""
        qdb.sql_connection.TRN.add(sql, [filepath_id])
        res = dict(qdb.sql_connection.TRN.execute_fetchindex()[0])

        obj_id = res.pop("artifact_id")
        res["fullpath"] = _path_builder(
            get_db_files_base_dir(),
            res["filepath"],
            res["mountpoint"],
            res["subdirectory"],
            obj_id,
        )
        return res


def filepath_id_to_rel_path(filepath_id):
    """Gets the relative to the base directory of filepath_id

    Returns
    -------
    str
        The relative path for the given filepath id
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT mountpoint, filepath, subdirectory, artifact_id
                 FROM qiita.filepath
                    JOIN qiita.data_directory USING (data_directory_id)
                    LEFT JOIN qiita.artifact_filepath USING (filepath_id)
                 WHERE filepath_id = %s"""
        qdb.sql_connection.TRN.add(sql, [filepath_id])
        # It should be only one row
        mp, fp, sd, a_id = qdb.sql_connection.TRN.execute_fetchindex()[0]
        if sd:
            result = join(mp, str(a_id), fp)
        else:
            result = join(mp, fp)
        return result


def filepath_id_to_object_id(filepath_id):
    """Gets the object id to which the filepath id belongs to

    Returns
    -------
    int
        The object id the filepath id belongs to or None if not found

    Notes
    -----
    This helper function is intended to be used with the download handler so
    we can prepend downloads with the artifact id; thus, we will only look for
    filepath ids in qiita.analysis_filepath and qiita.artifact_filepath as
    search in qiita.reference, qiita.prep_template_filepath and
    qiita.sample_template_filepath will make the naming redundat (those already
    have the study_id in their filename)
    """
    with qdb.sql_connection.TRN:
        sql = """
            SELECT analysis_id FROM qiita.analysis_filepath
                WHERE filepath_id = %s UNION
            SELECT artifact_id FROM qiita.artifact_filepath
                WHERE filepath_id = %s"""
        qdb.sql_connection.TRN.add(sql, [filepath_id, filepath_id])
        fids = sorted(qdb.sql_connection.TRN.execute_fetchflatten())
        if fids:
            return fids[0]
        return None


def filepath_ids_to_rel_paths(filepath_ids):
    """Gets the full paths, relative to the base directory

    Parameters
    ----------
    filepath_ids : list of int

    Returns
    -------
    dict where keys are ints and values are str
        {filepath_id: relative_path}
    """
    if not filepath_ids:
        return {}

    with qdb.sql_connection.TRN:
        sql = """SELECT filepath_id, mountpoint, filepath, subdirectory,
                        artifact_id
                 FROM qiita.filepath
                    JOIN qiita.data_directory USING (data_directory_id)
                    LEFT JOIN qiita.artifact_filepath USING (filepath_id)
                 WHERE filepath_id IN %s"""
        qdb.sql_connection.TRN.add(sql, [tuple(filepath_ids)])
        res = {}
        for row in qdb.sql_connection.TRN.execute_fetchindex():
            if row[3]:
                res[row[0]] = join(row[1], str(row[4]), row[2])
            else:
                res[row[0]] = join(row[1], row[2])
        return res


def convert_to_id(value, table, text_col=None):
    """Converts a string value to its corresponding table identifier

    Parameters
    ----------
    value : str
        The string value to convert
    table : str
        The table that has the conversion
    text_col : str, optional
        Column holding the string value. Defaults to same as table name.

    Returns
    -------
    int
        The id correspinding to the string

    Raises
    ------
    QiitaDBLookupError
        The passed string has no associated id
    """
    text_col = table if text_col is None else text_col
    with qdb.sql_connection.TRN:
        sql = "SELECT {0}_id FROM qiita.{0} WHERE {1} = %s".format(table, text_col)
        qdb.sql_connection.TRN.add(sql, [value])
        _id = qdb.sql_connection.TRN.execute_fetchindex()
        if not _id:
            raise qdb.exceptions.QiitaDBLookupError(
                "%s not valid for table %s" % (value, table)
            )
        # If there was a result it was a single row and and single value,
        # hence access to [0][0]
        return _id[0][0]


def convert_from_id(value, table):
    """Converts an id value to its corresponding string value

    Parameters
    ----------
    value : int
        The id value to convert
    table : str
        The table that has the conversion

    Returns
    -------
    str
        The string correspinding to the id

    Raises
    ------
    QiitaDBLookupError
        The passed id has no associated string
    """
    with qdb.sql_connection.TRN:
        sql = "SELECT {0} FROM qiita.{0} WHERE {0}_id = %s".format(table)
        qdb.sql_connection.TRN.add(sql, [value])
        string = qdb.sql_connection.TRN.execute_fetchindex()
        if not string:
            raise qdb.exceptions.QiitaDBLookupError(
                "%s not valid for table %s" % (value, table)
            )
        # If there was a result it was a single row and and single value,
        # hence access to [0][0]
        return string[0][0]


def get_count(table):
    """Counts the number of rows in a table

    Parameters
    ----------
    table : str
        The name of the table of which to count the rows

    Returns
    -------
    int
    """
    with qdb.sql_connection.TRN:
        sql = "SELECT count(1) FROM %s" % table
        qdb.sql_connection.TRN.add(sql)
        return qdb.sql_connection.TRN.execute_fetchlast()


def check_count(table, exp_count):
    """Checks that the number of rows in a table equals the expected count

    Parameters
    ----------
    table : str
        The name of the table of which to count the rows
    exp_count : int
        The expected number of rows in the table

    Returns
    -------
    bool
    """
    obs_count = get_count(table)
    return obs_count == exp_count


def get_environmental_packages():
    """Get the list of available environmental packages

    Returns
    -------
    list of (str, str)
        The available environmental packages. The first string is the
        environmental package name and the second string is the table where
        the metadata for the environmental package is stored
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT * FROM qiita.environmental_package")
        return qdb.sql_connection.TRN.execute_fetchindex()


def get_visibilities():
    """Get the list of available visibilities for artifacts

    Returns
    -------
    list of str
        The available visibilities
    """
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add("SELECT visibility FROM qiita.visibility")
        return qdb.sql_connection.TRN.execute_fetchflatten()


def get_timeseries_types():
    """Get the list of available timeseries types

    Returns
    -------
    list of (int, str, str)
        The available timeseries types. Each timeseries type is defined by the
        tuple (timeseries_id, timeseries_type, intervention_type)
    """
    with qdb.sql_connection.TRN:
        sql = "SELECT * FROM qiita.timeseries_type ORDER BY timeseries_type_id"
        qdb.sql_connection.TRN.add(sql)
        return qdb.sql_connection.TRN.execute_fetchindex()


def get_pubmed_ids_from_dois(doi_ids):
    """Get the dict of pubmed ids from a list of doi ids

    Parameters
    ----------
    doi_ids : list of str
        The list of doi ids

    Returns
    -------
    dict of {doi: pubmed_id}
        Return dict of doi and pubmed ids

    Notes
    -----
    If doi doesn't exist it will not return that {key: value} pair
    """
    with qdb.sql_connection.TRN:
        sql = "SELECT doi, pubmed_id FROM qiita.publication WHERE doi IN %s"
        qdb.sql_connection.TRN.add(sql, [tuple(doi_ids)])
        return {row[0]: row[1] for row in qdb.sql_connection.TRN.execute_fetchindex()}


def infer_status(statuses):
    """Infers an object status from the statuses passed in

    Parameters
    ----------
    statuses : list of lists of strings or empty list
        The list of statuses used to infer the resulting status (the result
        of execute_fetchall)

    Returns
    -------
    str
        The inferred status

    Notes
    -----
    The inference is done in the following priority (high to low):
        (1) public
        (2) private
        (3) awaiting_approval
        (4) sandbox
    """
    if statuses:
        statuses = set(s[0] for s in statuses)
        if "public" in statuses:
            return "public"
        if "private" in statuses:
            return "private"
        if "awaiting_approval" in statuses:
            return "awaiting_approval"
    # If there are no statuses, or any of the previous ones have been found
    # then the inferred status is 'sandbox'
    return "sandbox"


def add_message(message, users):
    """Adds a message to the messages table, attaching it to given users

    Parameters
    ----------
    message : str
        Message to add
    users : list of User objects
        Users to connect the message to
    """
    with qdb.sql_connection.TRN:
        sql = """INSERT INTO qiita.message (message) VALUES (%s)
                 RETURNING message_id"""
        qdb.sql_connection.TRN.add(sql, [message])
        msg_id = qdb.sql_connection.TRN.execute_fetchlast()
        sql = """INSERT INTO qiita.message_user (email, message_id)
                 VALUES (%s, %s)"""
        sql_args = [[user.id, msg_id] for user in users]
        qdb.sql_connection.TRN.add(sql, sql_args, many=True)
        qdb.sql_connection.TRN.execute()


def add_system_message(message, expires):
    """Adds a system message to the messages table, attaching it to asl users

    Parameters
    ----------
    message : str
        Message to add
    expires : datetime object
        Expiration for the message
    """
    with qdb.sql_connection.TRN:
        sql = """INSERT INTO qiita.message (message, expiration)
                 VALUES (%s, %s)
                 RETURNING message_id"""
        qdb.sql_connection.TRN.add(sql, [message, expires])
        msg_id = qdb.sql_connection.TRN.execute_fetchlast()
        sql = """INSERT INTO qiita.message_user (email, message_id)
                 SELECT email, %s FROM qiita.qiita_user"""
        qdb.sql_connection.TRN.add(sql, [msg_id])
        qdb.sql_connection.TRN.execute()


def clear_system_messages():
    with qdb.sql_connection.TRN:
        sql = "SELECT message_id FROM qiita.message WHERE expiration < %s"
        qdb.sql_connection.TRN.add(sql, [datetime.now()])
        msg_ids = qdb.sql_connection.TRN.execute_fetchflatten()
        if msg_ids:
            msg_ids = tuple(msg_ids)
            sql = "DELETE FROM qiita.message_user WHERE message_id IN %s"
            qdb.sql_connection.TRN.add(sql, [msg_ids])
            sql = "DELETE FROM qiita.message WHERE message_id IN %s"
            qdb.sql_connection.TRN.add(sql, [msg_ids])
            qdb.sql_connection.TRN.execute()


def supported_filepath_types(artifact_type):
    """Returns the list of supported filepath types for the given artifact type

    Parameters
    ----------
    artifact_type : str
        The artifact type to check the supported filepath types

    Returns
    -------
    list of [str, bool]
        The list of supported filepath types and whether it is required by the
        artifact type or not
    """
    with qdb.sql_connection.TRN:
        sql = """SELECT DISTINCT filepath_type, required
                 FROM qiita.artifact_type_filepath_type
                    JOIN qiita.artifact_type USING (artifact_type_id)
                    JOIN qiita.filepath_type USING (filepath_type_id)
                 WHERE artifact_type = %s"""
        qdb.sql_connection.TRN.add(sql, [artifact_type])
        return qdb.sql_connection.TRN.execute_fetchindex()


def generate_study_list(user, visibility):
    """Get general study information

    Parameters
    ----------
    user : qiita_db.user.User
        The user of which we are requesting studies from
    visibility : string
        The visibility to get studies {'public', 'user'}

    Returns
    -------
    list of dict
        The list of studies and their information

    Notes
    -----
    The main select might look scary but it's pretty simple:
    - We select the requiered fields from qiita.study and qiita.study_person
        SELECT metadata_complete, study_abstract, study_id, study_alias,
            study_title, ebi_study_accession, autoloaded,
            qiita.study_person.name AS pi_name,
            qiita.study_person.email AS pi_email,
    - the total number of samples collected by counting sample_ids
            (SELECT COUNT(sample_id) FROM qiita.study_sample
                WHERE study_id=qiita.study.study_id)
                AS number_samples_collected]
    - retrieve all the prep data types for all the artifacts depending on their
      visibility
            (SELECT array_agg(row_to_json((prep_template_id, data_type,
                 artifact_id, artifact_type, deprecated,
                 qiita.bioms_from_preparation_artifacts(prep_template_id)),
                 true))
                FROM qiita.study_prep_template
                LEFT JOIN qiita.prep_template USING (prep_template_id)
                LEFT JOIN qiita.data_type USING (data_type_id)
                LEFT JOIN qiita.artifact USING (artifact_id)
                LEFT JOIN qiita.artifact_type USING (artifact_type_id)
                LEFT JOIN qiita.visibility USING (visibility_id)
                WHERE {0} study_id = qiita.study.study_id)
                    AS preparation_information,
    - all the publications that belong to the study
            (SELECT array_agg((publication, is_doi)))
                FROM qiita.study_publication
                WHERE study_id=qiita.study.study_id) AS publications,
    - all names sorted by email of users that have access to the study
            (SELECT array_agg(name ORDER BY email) FROM qiita.study_users
                LEFT JOIN qiita.qiita_user USING (email)
                WHERE study_id=qiita.study.study_id) AS shared_with_name,
    - all emails sorted by email of users that have access to the study
            (SELECT array_agg(email ORDER BY email) FROM qiita.study_users
                LEFT JOIN qiita.qiita_user USING (email)
                WHERE study_id=qiita.study.study_id) AS shared_with_email
    - all study tags
            (SELECT array_agg(study_tag) FROM qiita.per_study_tags
                WHERE study_id=qiita.study.study_id) AS study_tags
    - study owner
            (SELECT name FROM qiita.qiita_user
                WHERE email=qiita.study.email) AS owner
    """

    visibility_sql = ""
    sids = set(s.id for s in user.user_studies.union(user.shared_studies))
    if visibility == "user":
        if user.level == "admin":
            sids = (
                sids
                | qdb.study.Study.get_ids_by_status("sandbox")
                | qdb.study.Study.get_ids_by_status("private")
                | qdb.study.Study.get_ids_by_status("awaiting_approval")
            )
    elif visibility == "public":
        sids = qdb.study.Study.get_ids_by_status("public") - sids
        visibility_sql = "visibility = 'public' AND"
    else:
        raise ValueError("Not a valid visibility: %s" % visibility)

    sql = """
        SELECT metadata_complete, study_abstract, study_id, study_alias,
            study_title, ebi_study_accession, autoloaded,
            qiita.study_person.name AS pi_name,
            qiita.study_person.email AS pi_email,
            (SELECT COUNT(sample_id) FROM qiita.study_sample
                WHERE study_id=qiita.study.study_id)
                AS number_samples_collected,
            (SELECT EXISTS(
                SELECT 1 FROM qiita.study_sample
                    WHERE study_id = qiita.study.study_id LIMIT 1))
                    AS has_sample_info,
            (SELECT array_agg(row_to_json((prep_template_id, data_type,
                 artifact_id, artifact_type, deprecated,
                 qiita.bioms_from_preparation_artifacts(prep_template_id)),
                 true))
                FROM qiita.study_prep_template
                LEFT JOIN qiita.prep_template USING (prep_template_id)
                LEFT JOIN qiita.data_type USING (data_type_id)
                LEFT JOIN qiita.artifact USING (artifact_id)
                LEFT JOIN qiita.artifact_type USING (artifact_type_id)
                LEFT JOIN qiita.visibility USING (visibility_id)
                WHERE {0} study_id = qiita.study.study_id)
                    AS preparation_information,
            (SELECT array_agg(row_to_json((publication, is_doi), true))
                FROM qiita.study_publication
                WHERE study_id=qiita.study.study_id) AS publications,
            (SELECT array_agg(name ORDER BY email) FROM qiita.study_users
                LEFT JOIN qiita.qiita_user USING (email)
                WHERE study_id=qiita.study.study_id) AS shared_with_name,
            (SELECT array_agg(email ORDER BY email) FROM qiita.study_users
                LEFT JOIN qiita.qiita_user USING (email)
                WHERE study_id=qiita.study.study_id) AS shared_with_email,
            (SELECT array_agg(study_tag) FROM qiita.per_study_tags
                WHERE study_id=qiita.study.study_id) AS study_tags,
            (SELECT name FROM qiita.qiita_user
                WHERE email=qiita.study.email) AS owner,
            qiita.study.email AS owner_email
            FROM qiita.study
            LEFT JOIN qiita.study_person ON (
                study_person_id=principal_investigator_id)
            WHERE study_id IN %s
            ORDER BY study_id""".format(visibility_sql)

    infolist = []
    if sids:
        with qdb.sql_connection.TRN:
            qdb.sql_connection.TRN.add(sql, [tuple(sids)])
            results = qdb.sql_connection.TRN.execute_fetchindex()

        for info in results:
            info = dict(info)

            # cleaning owners name
            if info["owner"] in (None, ""):
                info["owner"] = info["owner_email"]
            del info["owner_email"]

            preparation_data_types = []
            artifact_biom_ids = []
            if info["preparation_information"] is not None:
                for pinfo in info["preparation_information"]:
                    # 'f1': prep_template_id, 'f2': data_type,
                    # 'f3': artifact_id, 'f4': artifact_type,
                    # 'f5':deprecated, 'f6': biom artifacts
                    if pinfo["f5"]:
                        continue
                    preparation_data_types.append(pinfo["f2"])
                    if pinfo["f4"] == "BIOM":
                        artifact_biom_ids.append(pinfo["f3"])
                    if pinfo["f6"] is not None:
                        artifact_biom_ids.extend(map(int, pinfo["f6"].split(",")))
            del info["preparation_information"]
            info["artifact_biom_ids"] = list(set(artifact_biom_ids))
            info["preparation_data_types"] = list(set(preparation_data_types))

            # publication info
            info["publication_doi"] = []
            info["publication_pid"] = []
            if info["publications"] is not None:
                for p in info["publications"]:
                    # f1-2 are the default names given by pgsql
                    pub = p["f1"]
                    is_doi = p["f2"]
                    if is_doi:
                        info["publication_doi"].append(pub)
                    else:
                        info["publication_pid"].append(pub)
            del info["publications"]

            # pi info
            info["pi"] = (info["pi_email"], info["pi_name"])
            del info["pi_email"]
            del info["pi_name"]

            # shared with
            info["shared"] = []
            if info["shared_with_name"] and info["shared_with_email"]:
                for name, email in zip(
                    info["shared_with_name"], info["shared_with_email"]
                ):
                    if not name:
                        name = email
                    info["shared"].append((email, name))
            del info["shared_with_name"]
            del info["shared_with_email"]

            # # add extra info about sample information file
            # if info['has_sample_info']:
            #     # the fix for #3091 should go here; please reference that
            #     # issue for more information of why it hasn't been closed
            #     with qdb.sql_connection.TRN:
            #         # check if host_scientific_name is part of the metadata
            #         BMT = qdb.metadata_template.base_metadata_template
            #         QCN = BMT.QIITA_COLUMN_NAME
            #         sql = """SELECT POSITION('host_scientific_name' IN
            #                                  sample_values->>'columns')
            #                  FROM qiita.sample_%d
            #                  WHERE sample_id = '%s'""" % (
            #                     info['study_id'], QCN)
            #         qdb.sql_connection.TRN.add(sql)
            #         has_hsn = qdb.sql_connection.TRN.execute_fetchflatten()
            #         # if it has that column, we can retrieve the information
            #         if has_hsn[0] != 0:
            #             sql = """SELECT array_agg(
            #                         DISTINCT
            #                         sample_values->>'host_scientific_name')
            #                      FROM qiita.sample_%d
            #                      WHERE sample_id != '%s'""" % (
            #                         info['study_id'], QCN))
            #             qdb.sql_connection.TRN.add(sql)
            #             hsn = qdb.sql_connection.TRN.execute_fetchflatten()
            #             info['host_scientific_name'] = hsn
            del info["has_sample_info"]

            infolist.append(info)
    return infolist


def generate_study_list_without_artifacts(study_ids, portal=None):
    """Get general study information without artifacts

    Parameters
    ----------
    study_ids : list of ints
        The study ids to look for. Non-existing ids will be ignored
    portal : str
        Portal to use, if None take it from configuration. Mainly for tests.

    Returns
    -------
    list of dict
        The list of studies and their information

    Notes
    -----
    The main select might look scary but it's pretty simple:
    - We select the requiered fields from qiita.study and qiita.study_person
        SELECT metadata_complete, study_abstract, study_id, study_alias,
            study_title, ebi_study_accession, autoloaded,
            qiita.study_person.name AS pi_name,
            qiita.study_person.email AS pi_email,
    - the total number of samples collected by counting sample_ids
            (SELECT COUNT(sample_id) FROM qiita.study_sample
                WHERE study_id=qiita.study.study_id)
                AS number_samples_collected]
    - all the publications that belong to the study
            (SELECT array_agg((publication, is_doi)))
                FROM qiita.study_publication
                WHERE study_id=qiita.study.study_id) AS publications
    """
    if portal is None:
        portal = qiita_config.portal
    with qdb.sql_connection.TRN:
        sql = """
            SELECT metadata_complete, study_abstract, study_id, study_alias,
                study_title, ebi_study_accession, autoloaded,
                qiita.study_person.name AS pi_name,
                qiita.study_person.email AS pi_email,
                (SELECT COUNT(sample_id) FROM qiita.study_sample
                    WHERE study_id=qiita.study.study_id)
                    AS number_samples_collected,
                (SELECT array_agg(row_to_json((publication, is_doi), true))
                    FROM qiita.study_publication
                    WHERE study_id=qiita.study.study_id) AS publications
                FROM qiita.study
                LEFT JOIN qiita.study_portal USING (study_id)
                LEFT JOIN qiita.portal_type USING (portal_type_id)
                LEFT JOIN qiita.study_person ON (
                    study_person_id=principal_investigator_id)
                WHERE study_id IN %s AND portal = %s
                ORDER BY study_id"""
        qdb.sql_connection.TRN.add(sql, [tuple(study_ids), portal])
        infolist = []
        for info in qdb.sql_connection.TRN.execute_fetchindex():
            info = dict(info)

            # publication info
            info["publication_doi"] = []
            info["publication_pid"] = []
            if info["publications"] is not None:
                for p in info["publications"]:
                    # f1-2 are the default names given
                    pub = p["f1"]
                    is_doi = p["f2"]
                    if is_doi:
                        info["publication_doi"].append(pub)
                    else:
                        info["publication_pid"].append(pub)
            del info["publications"]

            # pi info
            info["pi"] = (info["pi_email"], info["pi_name"])
            del info["pi_email"]
            del info["pi_name"]

            infolist.append(info)
    return infolist


def get_artifacts_information(artifact_ids, only_biom=True):
    """Returns processing information about the artifact ids

    Parameters
    ----------
    artifact_ids : list of ints
        The artifact ids to look for. Non-existing ids will be ignored
    only_biom : bool
        If true only the biom artifacts are retrieved

    Returns
    -------
    dict
        The info of the artifacts
    """
    if not artifact_ids:
        return {}

    sql = """
        WITH main_query AS (
            SELECT a.artifact_id, a.name, a.command_id as cid, sc.name,
                   a.generated_timestamp, array_agg(a.command_parameters),
                   dt.data_type, parent_id,
                   parent_info.command_id, parent_info.name,
                   array_agg(parent_info.command_parameters),
                   array_agg(filepaths.filepath),
                   qiita.find_artifact_roots(a.artifact_id) AS root_id
            FROM qiita.artifact a
            LEFT JOIN qiita.software_command sc USING (command_id)"""
    if only_biom:
        sql += """
            JOIN qiita.artifact_type at ON (
                a.artifact_type_id = at .artifact_type_id
                    AND artifact_type = 'BIOM')"""
    sql += """
            LEFT JOIN qiita.parent_artifact pa ON (
                a.artifact_id = pa.artifact_id)
            LEFT JOIN qiita.data_type dt USING (data_type_id)
            LEFT OUTER JOIN LATERAL (
                SELECT command_id, sc.name, command_parameters
                FROM qiita.artifact ap
                LEFT JOIN qiita.software_command sc USING (command_id)
                WHERE ap.artifact_id = pa.parent_id) parent_info ON true
            LEFT OUTER JOIN LATERAL (
                SELECT filepath
                FROM qiita.artifact_filepath af
                JOIN qiita.filepath USING (filepath_id)
                WHERE af.artifact_id = a.artifact_id) filepaths ON true
            WHERE a.artifact_id IN %s
                AND a.visibility_id NOT IN %s
            GROUP BY a.artifact_id, a.name, a.command_id, sc.name,
                     a.generated_timestamp, dt.data_type, parent_id,
                     parent_info.command_id, parent_info.name
            ORDER BY a.command_id, artifact_id),
          has_target_subfragment AS (
            SELECT main_query.*, prep_template_id
            FROM main_query
            LEFT JOIN qiita.prep_template pt ON (
                main_query.root_id = pt.artifact_id)
        )
        SELECT * FROM has_target_subfragment
        ORDER BY cid, data_type, artifact_id
        """

    sql_params = """SELECT command_id, array_agg(parameter_name)
                    FROM qiita.command_parameter
                    WHERE parameter_type = 'artifact'
                    GROUP BY command_id"""

    QCN = qdb.metadata_template.base_metadata_template.QIITA_COLUMN_NAME
    sql_ts = """SELECT DISTINCT sample_values->>'target_subfragment'
                FROM qiita.prep_%s
                WHERE sample_id != '{0}'""".format(QCN)

    with qdb.sql_connection.TRN:
        results = []

        # getting all commands and their artifact parameters so we can
        # delete from the results below
        commands = {}
        qdb.sql_connection.TRN.add(sql_params)
        for cid, params in qdb.sql_connection.TRN.execute_fetchindex():
            cmd = qdb.software.Command(cid)
            commands[cid] = {
                "params": params,
                "merging_scheme": cmd.merging_scheme,
                "active": cmd.active,
                "deprecated": cmd.software.deprecated,
            }

        # Now let's get the actual artifacts. Note that ts is a cache
        # (prep id : target subfragment) so we don't have to query
        # multiple times the target subfragment for a prep info file.
        # However, some artifacts (like analysis) do not have a prep info
        # file; thus we can have a None prep id (key)
        ts = {None: []}
        ps = {}
        algorithm_az = {"": ""}
        PT = qdb.metadata_template.prep_template.PrepTemplate
        qdb.sql_connection.TRN.add(
            sql, [tuple(artifact_ids), qdb.util.artifact_visibilities_to_skip()]
        )
        for row in qdb.sql_connection.TRN.execute_fetchindex():
            (
                aid,
                name,
                cid,
                cname,
                gt,
                aparams,
                dt,
                pid,
                pcid,
                pname,
                pparams,
                filepaths,
                _,
                prep_template_id,
            ) = row

            # cleaning up aparams & pparams
            # - [0] due to the array_agg
            aparams = aparams[0]
            pparams = pparams[0]
            if aparams is None:
                aparams = {}
            else:
                # we are going to remove any artifacts from the parameters
                for ti in commands[cid]["params"]:
                    del aparams[ti]

            # - ignoring empty filepaths
            if filepaths == [None]:
                filepaths = []
            else:
                filepaths = [fp for fp in filepaths if fp.endswith("biom")]

            # generating algorithm, by default is ''
            algorithm = ""
            if cid is not None:
                deprecated = commands[cid]["deprecated"]
                active = commands[cid]["active"]
                if pcid is None:
                    parent_merging_scheme = None
                else:
                    parent_merging_scheme = commands[pcid]["merging_scheme"]

                algorithm = human_merging_scheme(
                    cname,
                    commands[cid]["merging_scheme"],
                    pname,
                    parent_merging_scheme,
                    aparams,
                    filepaths,
                    pparams,
                )

                if algorithm not in algorithm_az:
                    algorithm_az[algorithm] = hashlib.md5(
                        algorithm.encode("utf-8")
                    ).hexdigest()
            else:
                # there is no cid, thus is a direct upload; setting things
                # like this so the artifacts are dispayed
                deprecated = False
                active = True

            if prep_template_id not in ts:
                qdb.sql_connection.TRN.add(sql_ts, [prep_template_id])
                ts[prep_template_id] = qdb.sql_connection.TRN.execute_fetchflatten()
            target = ts[prep_template_id]

            prep_samples = 0
            platform = "not provided"
            target_gene = "not provided"
            if prep_template_id is not None:
                if prep_template_id not in ps:
                    pt = PT(prep_template_id)
                    categories = pt.categories
                    if "platform" in categories:
                        platform = ", ".join(set(pt.get_category("platform").values()))
                    if "target_gene" in categories:
                        target_gene = ", ".join(
                            set(pt.get_category("target_gene").values())
                        )

                    ps[prep_template_id] = [len(list(pt.keys())), platform, target_gene]

                prep_samples, platform, target_gene = ps[prep_template_id]

            results.append(
                {
                    "artifact_id": aid,
                    "target_subfragment": target,
                    "prep_samples": prep_samples,
                    "platform": platform,
                    "target_gene": target_gene,
                    "name": name,
                    "data_type": dt,
                    "timestamp": str(gt),
                    "parameters": aparams,
                    "algorithm": algorithm,
                    "algorithm_az": algorithm_az[algorithm],
                    "deprecated": deprecated,
                    "active": active,
                    "files": filepaths,
                }
            )

        return results


def _is_string_or_bytes(s):
    """Returns True if input argument is string (unicode or not) or bytes."""
    return isinstance(s, str) or isinstance(s, bytes)


def _get_filehandle(filepath_or, *args, **kwargs):
    """Open file if `filepath_or` looks like a string/unicode/bytes/Excel, else
    pass through.

    Notes
    -----
    If Excel, the code will write a temporary txt file with the contents. Also,
    it will check if the file is a Qiimp file or a regular Excel file.
    """
    if _is_string_or_bytes(filepath_or):
        if h5py.is_hdf5(filepath_or):
            fh, own_fh = h5py.File(filepath_or, *args, **kwargs), True
        elif filepath_or.endswith(".xlsx"):
            # due to extension, let's assume Excel file
            wb = load_workbook(filename=filepath_or, data_only=True)
            sheetnames = wb.sheetnames
            # let's check if Qiimp, they must be in same order
            first_cell_index = 0
            is_qiimp_wb = False
            if sheetnames == [
                "Metadata",
                "Validation",
                "Data Dictionary",
                "metadata_schema",
                "metadata_form",
                "Instructions",
            ]:
                first_cell_index = 1
                is_qiimp_wb = True
            first_sheet = wb[sheetnames[0]]
            cell_range = range(first_cell_index, first_sheet.max_column)
            _, fp = mkstemp(suffix=".txt")
            with open(fp, "w") as fh:
                cfh = csv_writer(fh, delimiter="\t")
                for r in first_sheet.rows:
                    if is_qiimp_wb:
                        # check contents of first column; if they are a zero
                        # (not a valid QIIMP sample_id) or a "No more than
                        # max samples" message, there are no more valid rows,
                        # so don't examine any more rows.
                        fcv = str(r[cell_range[0]].value)
                        if fcv == "0" or fcv.startswith("No more than"):
                            break
                    cfh.writerow([r[x].value for x in cell_range])
            fh, own_fh = open(fp, *args, **kwargs), True
        else:
            fh, own_fh = open(filepath_or, *args, **kwargs), True
    else:
        fh, own_fh = filepath_or, False
    return fh, own_fh


@contextmanager
def open_file(filepath_or, *args, **kwargs):
    """Context manager, like ``open``, but lets file handles and file like
    objects pass untouched.

    It is useful when implementing a function that can accept both
    strings and file-like objects (like numpy.loadtxt, etc).

    This method differs slightly from scikit-bio's implementation in that it
    handles HDF5 files appropriately.

    Parameters
    ----------
    filepath_or : str/bytes/unicode string or file-like
         If string, file to be opened using ``h5py.File`` if the file is an
         HDF5 file, otherwise builtin ``open`` will be used. If it is not a
         string, the object is just returned untouched.

    Other parameters
    ----------------
    args, kwargs : tuple, dict
        When `filepath_or` is a string, any extra arguments are passed
        on to the ``open`` builtin.
    """
    fh, own_fh = _get_filehandle(filepath_or, *args, **kwargs)
    try:
        yield fh
    finally:
        if own_fh:
            fh.close()


def artifact_visibilities_to_skip():
    return tuple([qdb.util.convert_to_id("archived", "visibility")])


def generate_analysis_list(analysis_ids, public_only=False):
    """Get general analysis information

    Parameters
    ----------
    analysis_ids : list of ints
        The analysis ids to look for. Non-existing ids will be ignored
    public_only : bool, optional
        If true, return only public analyses. Default: false.

    Returns
    -------
    list of dict
        The list of studies and their information
    """
    if not analysis_ids:
        return []

    sql = """
        SELECT analysis_id, a.name, a.description, a.timestamp, a.email,
            array_agg(DISTINCT artifact_id),
            array_agg(DISTINCT visibility),
            array_agg(DISTINCT CASE WHEN filepath_type = 'plain_text'
                      THEN filepath_id END)
        FROM qiita.analysis a
        LEFT JOIN qiita.analysis_artifact USING (analysis_id)
        LEFT JOIN qiita.artifact USING (artifact_id)
        LEFT JOIN qiita.visibility USING (visibility_id)
        LEFT JOIN qiita.analysis_filepath USING (analysis_id)
        LEFT JOIN qiita.filepath USING (filepath_id)
        LEFT JOIN qiita.filepath_type USING (filepath_type_id)
        WHERE dflt = false AND analysis_id IN %s
        GROUP BY analysis_id
        ORDER BY analysis_id"""

    with qdb.sql_connection.TRN:
        results = []

        qdb.sql_connection.TRN.add(sql, [tuple(analysis_ids)])
        for row in qdb.sql_connection.TRN.execute_fetchindex():
            aid, name, description, ts, owner, artifacts, av, mapping_files = row

            av = "public" if set(av) == {"public"} else "private"
            if av != "public" and public_only:
                continue

            if mapping_files == [None]:
                mapping_files = []
            else:
                mapping_files = [
                    (mid, get_filepath_information(mid)["fullpath"])
                    for mid in mapping_files
                    if mid is not None
                ]
            if artifacts == [None]:
                artifacts = []
            else:
                # making sure they are int so they don't break the GUI
                artifacts = [int(a) for a in artifacts if a is not None]

            results.append(
                {
                    "analysis_id": aid,
                    "name": name,
                    "description": description,
                    "timestamp": ts.strftime("%m/%d/%y %H:%M:%S"),
                    "visibility": av,
                    "artifacts": artifacts,
                    "owner": owner,
                    "mapping_files": mapping_files,
                }
            )

    return results


def generate_analyses_list_per_study(study_id):
    """Get study analyses and their preparations

    Parameters
    ----------
    study_id : int
        The study id

    Returns
    -------
    list of dict
        The available analyses and their general information
    """
    # for speed and SQL simplicity, we are going to split the search in two
    # queries: 1. analysis_sql: to find analyses associated with this study
    # and the artifacts used to generate the analyses; and 2. extra_sql: each
    # analysis details, including the artifacts (children) that belong to
    # the analysis.
    analysis_sql = """
        SELECT DISTINCT analysis_id, array_agg(DISTINCT artifact_id) AS aids
        FROM qiita.analysis_sample analysis_sample
        WHERE sample_id IN (SELECT sample_id
                            FROM qiita.study_sample
                            WHERE study_id = %s)
        GROUP BY analysis_id
        ORDER BY analysis_id
    """
    extra_sql = """
        SELECT analysis_id, analysis.name, analysis.email, analysis.dflt,
            array_agg(DISTINCT aa.artifact_id) FILTER (
                      WHERE aa.artifact_id IS NOT NULL) as artifact_ids,
            ARRAY(SELECT DISTINCT prep_template_id
                  FROM qiita.preparation_artifact
                  WHERE artifact_id IN %s) as prep_ids,
            array_agg(DISTINCT visibility.visibility) FILTER (
                    WHERE aa.artifact_id IS NOT NULL) as visibility
        FROM qiita.analysis analysis
        LEFT JOIN qiita.analysis_artifact aa USING (analysis_id)
        LEFT JOIN qiita.artifact artifact USING (artifact_id)
        LEFT JOIN qiita.visibility visibility USING (visibility_id)
        WHERE analysis_id = %s
        GROUP BY analysis_id, analysis.name, analysis.email, analysis.dflt
    """
    results = []
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add(analysis_sql, [study_id])
        aids = qdb.sql_connection.TRN.execute_fetchindex()
        for aid, artifact_ids in aids:
            qdb.sql_connection.TRN.add(extra_sql, [tuple(artifact_ids), aid])
            for row in qdb.sql_connection.TRN.execute_fetchindex():
                results.append(dict(row))

    return results


def create_nested_path(path):
    """Wraps makedirs() to make it safe across multiple concurrent calls.
    Returns successfully if the path was created, or if it already exists.
    (Note, this alters the normal makedirs() behavior, where False is returned
    if the full path already exists.)

    Parameters
    ----------
    path : str
        The path to be created. The path can contain multiple levels that do
        not currently exist on the filesystem.

    Raises
    ------
    OSError
        If the operation failed for whatever reason (likely because the caller
        does not have permission to create new directories in the part of the
        filesystem requested
    """
    # TODO: catching errno=EEXIST (17 usually) will suffice for now, to avoid
    # stomping when multiple artifacts are being manipulated within a study.
    # In the future, employ a process-spanning mutex to serialize.
    # With Python3, the try/except wrapper can be replaced with a call to
    # makedirs with exist_ok=True
    try:
        # try creating the directory specified. if the directory already exists
        # , or if qiita does not have permissions to create/modify the path, an
        # exception will be thrown.
        makedirs(path)
    except OSError as e:
        # if the directory already exists, treat as success (idempotent)
        if e.errno != EEXIST:
            raise


def human_merging_scheme(
    cname,
    merging_scheme,
    pname,
    parent_merging_scheme,
    artifact_parameters,
    artifact_filepaths,
    parent_parameters,
):
    """From the artifact and its parent features format the merging scheme

    Parameters
    ----------
    cname : str
        The artifact command name
    merging_scheme : dict, from qdb.artifact.Artifact.merging_scheme
        The artifact merging scheme
    pname : str
        The artifact parent command name
    parent_merging_scheme : dict, from qdb.artifact.Artifact.merging_scheme
        The artifact parent merging scheme
    artifact_parameters : dict
        The artfiact processing parameters
    artifact_filepaths : list of str
        The artifact filepaths
    parent_parameters :
        The artifact parents processing parameters

    Returns
    -------
    str
        The merging scheme
    """
    eparams = []
    if merging_scheme["parameters"]:
        eparams.append(
            ",".join(
                [
                    "%s: %s" % (k, artifact_parameters[k])
                    for k in merging_scheme["parameters"]
                ]
            )
        )
    if (
        merging_scheme["outputs"]
        and artifact_filepaths is not None
        and artifact_filepaths
    ):
        eparams.append("BIOM: %s" % ", ".join(artifact_filepaths))
    if eparams:
        cname = "%s (%s)" % (cname, ", ".join(eparams))

    if merging_scheme["ignore_parent_command"]:
        algorithm = cname
    else:
        palgorithm = "N/A"
        if pname is not None:
            palgorithm = pname
            if parent_merging_scheme["parameters"]:
                params = ",".join(
                    [
                        "%s: %s" % (k, parent_parameters[k])
                        for k in parent_merging_scheme["parameters"]
                    ]
                )
                palgorithm = "%s (%s)" % (palgorithm, params)

        algorithm = "%s | %s" % (cname, palgorithm)

    return algorithm


def activate_or_update_plugins(update=False):
    """Activates/updates the plugins

    Parameters
    ----------
    update : bool, optional
        If True will update the plugins. Otherwise will activate them.
        Default: False.
    """
    conf_files = sorted(glob(join(qiita_config.plugin_dir, "*.conf")))
    label = "{} plugin (%s/{}): %s... ".format(
        "Updating" if update else "\tLoading", len(conf_files)
    )
    for i, fp in enumerate(conf_files):
        print(label % (i + 1, basename(fp)), end=None)
        s = qdb.software.Software.from_file(fp, update=update)
        if not update:
            s.activate()
        print("Ok")


def send_email(to, subject, body):
    # create email
    msg = MIMEMultipart()
    msg["From"] = qiita_config.smtp_email
    msg["To"] = to
    # we need to do 'replace' because the subject can have
    # new lines in the middle of the string
    msg["Subject"] = subject.replace("\n", "")
    msg.attach(MIMEText(body, "plain"))

    # connect to smtp server, using ssl if needed
    if qiita_config.smtp_ssl:
        smtp = SMTP_SSL()
    else:
        smtp = SMTP()
    smtp.set_debuglevel(False)
    smtp.connect(qiita_config.smtp_host, qiita_config.smtp_port)
    # try tls, if not available on server just ignore error
    try:
        smtp.starttls()
    except SMTPException:
        pass
    smtp.ehlo_or_helo_if_needed()

    if qiita_config.smtp_user:
        smtp.login(qiita_config.smtp_user, qiita_config.smtp_password)

    # send email
    try:
        smtp.sendmail(qiita_config.smtp_email, to, msg.as_string())
    except Exception:
        raise RuntimeError("Can't send email!")
    finally:
        smtp.close()


def resource_allocation_plot(df, col_name):
    """Builds resource allocation plot for given filename and jobs

    Parameters
    ----------
    file : str, required
        Builds plot for the specified file name. Usually provided as tsv.gz
    col_name: str, required
        Specifies x axis for the graph

    Returns
    ----------
    matplotlib.pyplot object
        Returns a matplotlib object with a plot
    """

    df.dropna(subset=["samples", "columns"], inplace=True)
    df[col_name] = df.samples * df["columns"]
    df[col_name] = df[col_name].astype(int)

    fig, axs = plt.subplots(ncols=2, figsize=(10, 4), sharey=False)

    ax = axs[0]
    mem_models, time_models = retrieve_equations()

    # models for memory
    _resource_allocation_plot_helper(df, ax, "MaxRSSRaw", mem_models, col_name)
    ax = axs[1]
    # models for time
    _resource_allocation_plot_helper(df, ax, "ElapsedRaw", time_models, col_name)

    return fig, axs


def retrieve_equations():
    """
    Helper function for resource_allocation_plot.
    Retrieves equations from db. Creates dictionary for memory and time models.

    Returns
    -------
    tuple
        dict
            memory models - potential memory models for resource allocations
        dict
            time models - potential time models for resource allocations
    """
    memory_models = {}
    time_models = {}
    res = []
    with qdb.sql_connection.TRN:
        sql = """ SELECT * FROM qiita.allocation_equations; """
        qdb.sql_connection.TRN.add(sql)
        res = qdb.sql_connection.TRN.execute_fetchindex()
    for models in res:
        if "mem" in models[1]:
            memory_models[models[1]] = {
                "equation_name": models[2],
                "equation": lambda x, k, a, b: eval(models[2]),
            }
        else:
            time_models[models[1]] = {
                "equation_name": models[2],
                "equation": lambda x, k, a, b: eval(models[2]),
            }
    return (memory_models, time_models)


def retrieve_resource_data(cname, sname, version, columns):
    """
    Retrieves resource data from db and constructs a DataFrame with relevant
    fields.

    Parameters
    ----------
    cname - command name for which we retrieve the resources
    sname - software name for which we retrieve the resources
    version - version of sftware for whhich we retrieve the resources
    columns - column names for the DataFrame returned by this function

    Returns
    -------
    pd.DataFrame
        DataFrame with resources.
    """
    with qdb.sql_connection.TRN:
        sql = """
            SELECT
                s.name AS sName,
                s.version AS sVersion,
                sc.command_id AS cID,
                sc.name AS cName,
                pr.processing_job_id AS processing_job_id,
                pr.command_parameters AS parameters,
                sra.samples AS samples,
                sra.columns AS columns,
                sra.input_size AS input_size,
                sra.extra_info AS extra_info,
                sra.memory_used AS memory_used,
                sra.walltime_used AS walltime_used,
                sra.job_start AS job_start,
                sra.node_name AS node_name,
                sra.node_model AS node_model
            FROM
                qiita.processing_job pr
            JOIN
                qiita.software_command sc ON pr.command_id = sc.command_id
            JOIN
                qiita.software s ON sc.software_id = s.software_id
            JOIN
                qiita.slurm_resource_allocations sra
                ON pr.processing_job_id = sra.processing_job_id
            WHERE
                sc.name = %s
                AND s.name = %s
                AND s.version = %s
            """
        qdb.sql_connection.TRN.add(sql, sql_args=[cname, sname, version])
        res = qdb.sql_connection.TRN.execute_fetchindex()
        df = pd.DataFrame(res, columns=columns)
        return df


def _resource_allocation_plot_helper(df, ax, curr, models, col_name):
    """Helper function for resource allocation plot. Builds plot for MaxRSSRaw
    and ElapsedRaw

    Parameters
    ----------
    df: pandas dataframe, required
        Filtered dataframe for the plot
    ax : matplotlib axes, required
        Axes for current subplot
    cname: str, required
        Specified job type
    sname: str, required
        Specified job sub type.
    col_name: str, required
        Specifies x axis for the graph
    curr: str, required
        Either MaxRSSRaw or ElapsedRaw (y axis)
    models: dictionary, required. Follows this structure
        equation_name: string
            Human readable representation of the equation
        equation: Python lambda function
            Lambda function representing equation to optimizse

    Returns
    -------
    best_model_name: string
        the name of the best model from the table
    best_model: function
        best fitting function for the current dictionary models
    options: object
        object containing constants for the best model (e.g. k, a, b in kx+b*a)
    """

    x_data, y_data = df[col_name], df[curr]
    # ax.scatter(x_data, y_data, s=2, label="data")
    d = dict()
    for index, row in df.iterrows():
        x_value = row[col_name]
        y_value = row[curr]
        if x_value not in d:
            d[x_value] = []
        d[x_value].append(y_value)

    for key in d.keys():
        # save only top point increased by 5% because our graph needs to exceed
        # the points
        d[key] = [max(d[key]) * 1.05]

    x_data = []
    y_data = []

    # Populate the lists with data from the dictionary
    for x, ys in d.items():
        for y in ys:
            x_data.append(x)
            y_data.append(y)

    x_data = np.array(x_data)
    y_data = np.array(y_data)
    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_ylabel(curr)
    ax.set_xlabel(col_name)

    # 50 - number of maximum iterations, 3 - number of failures we tolerate
    best_model_name, best_model, options = _resource_allocation_calculate(
        df, x_data, y_data, models, curr, col_name, 50, 3
    )
    k, a, b = options.x
    x_plot = np.array(sorted(df[col_name].unique()))
    y_plot = best_model(x_plot, k, a, b)
    ax.plot(x_plot, y_plot, linewidth=1, color="orange")

    cmin_value = min(y_plot)
    cmax_value = max(y_plot)

    maxi = (
        naturalsize(df[curr].max(), gnu=True)
        if curr == "MaxRSSRaw"
        else timedelta(seconds=float(df[curr].max()))
    )
    cmax = (
        naturalsize(cmax_value, gnu=True)
        if curr == "MaxRSSRaw"
        else str(timedelta(seconds=round(cmax_value, 2))).rstrip("0").rstrip(".")
    )

    mini = (
        naturalsize(df[curr].min(), gnu=True)
        if curr == "MaxRSSRaw"
        else timedelta(seconds=float(df[curr].min()))
    )
    cmin = (
        naturalsize(cmin_value, gnu=True)
        if curr == "MaxRSSRaw"
        else str(timedelta(seconds=round(cmin_value, 2))).rstrip("0").rstrip(".")
    )

    x_plot = np.array(df[col_name])
    success_df, failures_df = _resource_allocation_success_failures(
        df, k, a, b, best_model, col_name, curr
    )
    failures = failures_df.shape[0]
    ax.scatter(
        failures_df[col_name], failures_df[curr], color="red", s=3, label="failures"
    )
    success_df["node_name"] = success_df["node_name"].fillna("unknown")
    slurm_hosts = set(success_df["node_name"].tolist())
    cmap = colormaps.get_cmap("Accent")
    if len(slurm_hosts) > len(cmap.colors):
        raise ValueError(f"""'Accent' colormap only has {len(cmap.colors)}
                     colors, but {len(slurm_hosts)} hosts are provided.""")
    colors = cmap.colors[: len(slurm_hosts)]

    for i, host in enumerate(slurm_hosts):
        host_df = success_df[success_df["node_name"] == host]
        ax.scatter(host_df[col_name], host_df[curr], color=colors[i], s=3, label=host)
    ax.set_title(
        f"k||a||b: {k}||{a}||{b}\n"
        f"model: {models[best_model_name]['equation_name']}\n"
        f"real: {mini} || {maxi}\n"
        f"calculated: {cmin} || {cmax}\n"
        f"failures: {failures}"
    )
    ax.legend(loc="upper left")
    return best_model_name, best_model, options


def _resource_allocation_calculate(df, x, y, models, type_, col_name, depth, tolerance):
    """Helper function for resource allocation plot. Calculates best_model and
    best_result given the models list and x,y data.

    Parameters
    ----------
    x: pandas.Series (pandas column), required
        Represents x data for the function calculation
    y: pandas.Series (pandas column), required
        Represents y data for the function calculation
    type_: str, required
        current type (e.g. MaxRSSRaw)
    col_name: str, required
        Specifies x axis for the graph
    models: dictionary, required. Follows this structure
        equation_name: string
            Human readable representation of the equation
        equation: Python lambda function
            Lambda function representing equation to optimizse
    depth: int, required
        Maximum number of iterations in binary search
    tolerance: int, required,
        Tolerance to number of failures possible to be considered as a model

    Returns
    ----------
    best_model_name: string
        the name of the best model from the table
    best_model: function
        best fitting function for the current dictionary models
    best_result: object
        object containing constants for the best model (e.g. k, a, b in kx+b*a)
    """

    init = [1, 1, 1]
    best_model_name = None
    best_model = None
    best_result = None
    best_failures = np.inf
    best_max = np.inf
    for model_name, model in models.items():
        model_equation = model["equation"]
        # start values for binary search, where sl is left, sr is right
        # penalty weight must be positive & non-zero, hence, sl >= 1.
        # the upper bound for error can be an arbitrary large number
        sl = 1
        sr = 100000
        left = sl
        right = sr
        prev_failures = np.inf
        min_max = np.inf
        cnt = 0
        res = [1, 1, 1]  # k, a, b

        # binary search where we find the minimum penalty weight given the
        # scoring constraints defined in if/else statements.
        while left < right and cnt < depth:
            middle = (left + right) // 2
            options = minimize(
                _resource_allocation_custom_loss,
                init,
                args=(x, y, model_equation, middle),
            )
            k, a, b = options.x
            # important: here we take the 2nd (last) value of tuple since
            # the helper function returns success, then failures.
            failures_df = _resource_allocation_success_failures(
                df, k, a, b, model_equation, col_name, type_
            )[-1]
            y_plot = model_equation(x, k, a, b)
            if not any(y_plot):
                continue
            cmax = max(y_plot)
            cmin = min(y_plot)
            failures = failures_df.shape[0]

            if failures < prev_failures:
                prev_failures = failures
                right = middle
                min_max = cmax
                res = options

            elif failures > prev_failures:
                left = middle
            else:
                if cmin < 0:
                    left = middle
                elif cmax < min_max:
                    min_max = cmax
                    res = options
                    right = middle
                else:
                    right = middle

            # proceed with binary search in a window 10k to the right
            if left >= right and cnt < depth:
                sl += 10000
                sr += 10000
                left = sl
                right = sr

            cnt += 1

        # check whether we tolerate a couple failures
        # this is helpful if the model that has e.g. 1 failure is a better fit
        # overall based on maximum calculated value.
        is_acceptable_based_on_failures = (
            prev_failures <= tolerance
            or abs(prev_failures - best_failures) < tolerance
            or best_failures == np.inf
        )

        # case where less failures
        if is_acceptable_based_on_failures:
            if min_max <= best_max:
                best_failures = prev_failures
                best_max = min_max
                best_model_name = model_name
                best_model = model_equation
                best_result = res
    return best_model_name, best_model, best_result


def _resource_allocation_custom_loss(params, x, y, model, p):
    """Helper function for resource allocation plot. Calculates custom loss
    for given model.

    Parameters
    ----------
    params: list, required
        Initial list of integers for the given model
    x: pandas.Series (pandas column), required
        Represents x data for the function calculation
    y: pandas.Series (pandas column), required
        Represents y data for the function calculation
    model: Python function
        Lambda function representing current equation
    p: int, required
        Penalty weight for custom loss function

    Returns
    ----------
    float
        The mean of the list returned by the loss calculation (np.where)
    """
    k, a, b = params

    residuals = y - model(x, k, a, b)
    # Apply a heavier penalty to points below the curve
    penalty = p
    weighted_residuals = np.where(residuals > 0, penalty * residuals**2, residuals**2)
    return np.mean(weighted_residuals)


def _resource_allocation_success_failures(df, k, a, b, model, col_name, type_):
    """Helper function for resource allocation plot. Creates a dataframe with
    successes and failures given current model.

    Parameters
    ----------
    df: pandas.Dataframe, required
        Represents dataframe containing current jobs data
    k: int, required
        k constant in a model
    a: int, required
        a constant in a model
    b: int, required
        b constant in a model
    model: function, required
        Current function
    col_name: str, required
        Specifies x axis for the graph
    type_: str, required
        Specifies for which type we're getting failures (e.g. MaxRSSRaw)

    Returns
    ----------
    tuple with:
        pandas.Dataframe
            Dataframe containing successes for current type.
        pandas.Dataframe
            Dataframe containing failures for current type.
    """

    x_plot = np.array(df[col_name])
    df[f"c{type_}"] = model(x_plot, k, a, b)
    success_df = df[df[type_] <= df[f"c{type_}"]]
    failures_df = df[df[type_] > df[f"c{type_}"]]
    return (success_df, failures_df)


def MaxRSS_helper(x):
    if x[-1] == "K":
        y = float(x[:-1]) * 1000
    elif x[-1] == "M":
        y = float(x[:-1]) * 1000000
    elif x[-1] == "G":
        y = float(x[:-1]) * 1000000000
    else:
        y = float(x)
    return y


def update_resource_allocation_table(weeks=1, test=None):
    # Thu, Apr 27, 2023 old allocations (from barnacle) were changed to a
    # better allocation so we default start time 2023-04-28 to
    # use the latests for the newest version
    """
    Updates qiita.slurm_resource_allocation SQL table with jobs from slurm.
    Retrieves the most recent job available in the table and appends with
    the data.

    Parameters:
    ----------
    weeks: integer, optional
        Number of weeks for which we want to make a request from slurm.
    test: pandas.DataFrame, optional
        Represents dataframe containing slurm data from 2023-04-28. Used
        for testing only.
    """

    # retrieve the most recent timestamp
    sql_timestamp = """
            SELECT
                pj.external_job_id,
                sra.job_start
            FROM
                qiita.processing_job pj
            JOIN
                qiita.slurm_resource_allocations sra
            ON
                pj.processing_job_id = sra.processing_job_id
            ORDER BY
                sra.job_start DESC
            LIMIT 1;
        """

    dates = ["", ""]

    slurm_external_id = 0
    start_date = datetime.strptime("2023-04-28", "%Y-%m-%d")
    with qdb.sql_connection.TRN:
        sql = sql_timestamp
        qdb.sql_connection.TRN.add(sql)
        res = qdb.sql_connection.TRN.execute_fetchindex()
        if res:
            sei, sd = res[0]
            if sei is not None:
                slurm_external_id = sei
            if sd is not None:
                start_date = sd
        dates = [start_date, start_date + timedelta(weeks=weeks)]

    sql_command = """
            SELECT
                pj.processing_job_id AS processing_job_id,
                pj.external_job_id AS external_job_id
            FROM
                qiita.software_command sc
            JOIN
                qiita.processing_job pj ON pj.command_id = sc.command_id
            JOIN
                qiita.processing_job_status pjs
                ON pj.processing_job_status_id = pjs.processing_job_status_id
            LEFT JOIN
                qiita.slurm_resource_allocations sra
                ON pj.processing_job_id = sra.processing_job_id
            WHERE
                pjs.processing_job_status = 'success'
            AND
                pj.external_job_id ~ '^[0-9]+$'
            AND
                CAST(pj.external_job_id AS INTEGER) > %s
            AND
                sra.processing_job_id IS NULL;
        """
    df = pd.DataFrame()
    with qdb.sql_connection.TRN:
        qdb.sql_connection.TRN.add(sql_command, sql_args=[slurm_external_id])
        res = qdb.sql_connection.TRN.execute_fetchindex()
        df = pd.DataFrame(res, columns=["processing_job_id", "external_id"])
        df["external_id"] = df["external_id"].astype(int)

    data = []
    sacct = [
        "sacct",
        "-p",
        "--format=JobID,ElapsedRaw,MaxRSS,Submit,Start,End,CPUTimeRAW,"
        "ReqMem,AllocCPUs,AveVMSize,MaxVMSizeNode",
        "--starttime",
        dates[0].strftime("%Y-%m-%d"),
        "--endtime",
        dates[1].strftime("%Y-%m-%d"),
        "--user",
        "qiita",
        "--state",
        "CD",
    ]

    if test is not None:
        slurm_data = test
    else:
        rvals = check_output(sacct).decode("ascii")
        slurm_data = pd.read_csv(StringIO(rvals), sep="|")

    # In slurm, each JobID is represented by 3 rows in the dataframe:
    # - external_id:        overall container for the job and its associated
    #                       requests. When the Timelimit is hit, the container
    #                       would take care of completing/stopping the
    #                       external_id.batch job.
    # - external_id.batch:  it's a container job, it provides how
    #                       much memory it uses and cpus allocated, etc.
    # - external_id.extern: takes into account anything that happens
    #                       outside processing but yet is included in
    #                       the container resources. As in, if you ssh
    #                       to the node and do something additional or run
    #                       a prolog script, that processing would be under
    #                       external_id but separate from external_id.batch
    # Here we are going to merge all this info into a single row + some
    # other columns

    def merge_rows(rows):
        date_fmt = "%Y-%m-%dT%H:%M:%S"
        wait_time = datetime.strptime(
            rows.iloc[0]["Start"], date_fmt
        ) - datetime.strptime(rows.iloc[0]["Submit"], date_fmt)
        if rows.shape[0] >= 2:
            tmp = rows.iloc[1].copy()
        else:
            tmp = rows.iloc[0].copy()
        tmp["WaitTime"] = wait_time
        return tmp

    slurm_data["external_id"] = slurm_data["JobID"].apply(
        lambda x: int(x.split(".")[0])
    )
    slurm_data["external_id"] = slurm_data["external_id"].ffill()

    slurm_data = (
        slurm_data.groupby("external_id").apply(merge_rows).reset_index(drop=True)
    )

    # filter to only those jobs that are within the slurm_data df.
    eids = set(slurm_data["external_id"])
    df = df[df["external_id"].isin(eids)]

    for index, row in df.iterrows():
        job = qdb.processing_job.ProcessingJob(row["processing_job_id"])
        extra_info = ""
        eid = job.external_id

        cmd = job.command
        s = job.command.software
        try:
            samples, columns, input_size = job.shape
        except qdb.exceptions.QiitaDBUnknownIDError:
            # this will be raised if the study or the analysis has been
            # deleted; in other words, the processing_job was ran but the
            # details about it were erased when the user deleted them -
            # however, we keep the job for the record
            continue
        except TypeError as e:
            # similar to the except above, exept that for these 2 commands, we
            # have the study_id as None
            if cmd.name in {
                "create_sample_template",
                "delete_sample_template",
                "list_remote_files",
            }:
                continue
            else:
                raise e
        sname = s.name

        if cmd.name == "release_validators":
            ej = qdb.processing_job.ProcessingJob(job.parameters.values["job"])
            extra_info = ej.command.name
            samples, columns, input_size = ej.shape
        elif cmd.name == "complete_job":
            artifacts = loads(job.parameters.values["payload"])["artifacts"]
            if artifacts is not None:
                extra_info = ",".join(
                    {
                        x["artifact_type"]
                        for x in artifacts.values()
                        if "artifact_type" in x
                    }
                )
        elif cmd.name == "Validate":
            input_size = sum(
                [len(x) for x in loads(job.parameters.values["files"]).values()]
            )
            sname = f"{sname} - {job.parameters.values['artifact_type']}"
        elif cmd.name == "Alpha rarefaction curves [alpha_rarefaction]":
            extra_info = job.parameters.values[
                (
                    "The number of rarefaction depths to include between "
                    "min_depth and max_depth. (steps)"
                )
            ]
        curr = slurm_data[slurm_data["external_id"] == int(eid)].iloc[0]
        barnacle_info = curr["MaxVMSizeNode"]
        if len(barnacle_info) == 0:
            barnacle_info = [None, None]
        else:
            barnacle_info = barnacle_info.split("-")

        row_dict = {
            "processing_job_id": job.id,
            "samples": samples,
            "columns": columns,
            "input_size": input_size,
            "extra_info": extra_info,
            "ElapsedRaw": curr["ElapsedRaw"],
            "MaxRSS": curr["MaxRSS"],
            "Start": curr["Start"],
            "node_name": barnacle_info[0],
            "node_model": barnacle_info[1],
        }
        data.append(row_dict)
    df = pd.DataFrame(data)

    # This is important as we are transforming the MaxRSS to raw value
    # so we need to confirm that there is no other suffixes
    print("Make sure that only 0/K/M exist", set(df.MaxRSS.apply(lambda x: str(x)[-1])))

    # Generating new columns
    df["MaxRSSRaw"] = df.MaxRSS.apply(lambda x: MaxRSS_helper(str(x)))
    df["ElapsedRawTime"] = df.ElapsedRaw.apply(lambda x: timedelta(seconds=float(x)))
    df.replace({np.nan: None}, inplace=True)

    for index, row in df.iterrows():
        with qdb.sql_connection.TRN:
            sql = """
                INSERT INTO qiita.slurm_resource_allocations (
                    processing_job_id,
                    samples,
                    columns,
                    input_size,
                    extra_info,
                    memory_used,
                    walltime_used,
                    job_start,
                    node_name,
                    node_model
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            to_insert = [
                row["processing_job_id"],
                row["samples"],
                row["columns"],
                row["input_size"],
                row["extra_info"],
                row["MaxRSSRaw"],
                row["ElapsedRaw"],
                row["Start"],
                row["node_name"],
                row["node_model"],
            ]
            qdb.sql_connection.TRN.add(sql, sql_args=to_insert)
            qdb.sql_connection.TRN.execute()


def merge_overlapping_strings(str1, str2):
    """Helper function to merge 2 overlapping strings

    Parameters
    ----------
    str1: str
        Initial string
    str2: str
        End string

    Returns
    ----------
    str
        The merged strings
    """
    overlap = ""
    for i in range(1, min(len(str1), len(str2)) + 1):
        if str1.endswith(str2[:i]):
            overlap = str2[:i]
    return str1 + str2[len(overlap) :]
